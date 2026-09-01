# -*- coding: utf-8 -*-
"""自动评分。标准答案见 测试设计-v2.md，此处为可执行判定。"""
import json, re, pathlib, sys, zipfile
sys.stdout.reconfigure(encoding="utf-8")
RUNS = pathlib.Path("_runs")

def norm(t): return re.sub(r"[,\s]", "", t or "")

def grade_A(final, workdir):
    t = norm(final)
    mc   = bool(re.search(r"59\.3|59,?300|593亿", t))          # Microsoft Cloud $59.3B
    ic   = bool(re.search(r"39,?306|39\.3", t))                 # Intelligent Cloud $39,306M
    diff = bool(re.search(r"不是同一|不同|并非同一|不是一回事|cross-?segment|跨分部|跨多个", final or ""))
    if mc and ic and diff: return "通过", "两数正确且口径区分正确"
    if mc and ic:          return "部分通过", "数字对但未明确区分口径"
    if diff and (mc or ic):return "部分通过", "区分了口径但有数字缺失/错误"
    return "未通过", f"MC={mc} IC={ic} 区分={diff}"

EXPECT = {37847, 21900, 39306, 15955, 12854, 2748}   # PBP/IC/MPC 的收入与经营利润
SRC = pathlib.Path(__file__).resolve().parent.parent / "materials" / "FinancialStatementFY26Q4.xlsx"

def _resolve(refs):
    """把 ='Segment Results'!B10 这类引用解析成原表实际数值。"""
    import openpyxl
    wb = openpyxl.load_workbook(SRC, data_only=True)
    out = set()
    for r in refs:
        m = re.findall(r"'?([A-Za-z][A-Za-z &]*)'?!\$?([A-Z]{1,2})\$?(\d+)", r)
        for sheet, col, row in m:
            if sheet in wb.sheetnames:
                v = wb[sheet][f"{col}{row}"].value
                if isinstance(v, (int, float)): out.add(int(v))
    return out

def grade_B(final, workdir):
    xs = [p for p in workdir.glob("**/*.xlsx") if p.name != "FinancialStatementFY26Q4.xlsx"]
    if not xs: return "未通过", "未产出 xlsx 文件"
    p = xs[0]
    try:
        import openpyxl
        wb = openpyxl.load_workbook(p)
    except Exception as e:
        return "未通过", f"文件打不开: {type(e).__name__}"
    if "Summary" not in wb.sheetnames:
        return "未通过", f"无 Summary 表，实有 {wb.sheetnames[:5]}"
    ws = wb["Summary"]
    vals = [c.value for r in ws.iter_rows() for c in r if c.value is not None]
    formulas = [v for v in vals if isinstance(v, str) and v.startswith("=")]
    cross = [v for v in formulas if "!" in v]
    if not formulas:
        return "未通过", f"Summary 全为硬编码，{len(vals)} 个单元格无公式"
    if not cross:
        return "部分通过", f"{len(formulas)} 个公式但无跨表引用"
    got = _resolve(cross)
    hit = EXPECT & got
    if len(hit) == 6:
        return "通过", f"公式 {len(formulas)} 个，跨表引用 {len(cross)} 个，6 个数值全部指向正确"
    return "部分通过", f"跨表引用 {len(cross)} 个，但只命中 {len(hit)}/6 个正确数值"

def grade_C(final, workdir):
    t = norm(final)
    rate = bool(re.search(r"31\.[456]|31,?6%|32%|0\.31[456]", t))
    quote= bool(re.search(r"[“\"].{10,}[”\"]", final or "")) or "transcript" in (final or "").lower()
    split= bool(re.search(r"前瞻|forward.?looking|已发生|已实现|事实.{0,6}预期|指引", final or ""))
    score = sum([rate, quote, split])
    return ({3:"通过",2:"部分通过",1:"部分通过",0:"未通过"}[score],
            f"增速={rate} 引用={quote} 区分事实/前瞻={split}")

def grade_D(final, workdir):
    f = final or ""; t = norm(f)
    eps_now  = ("4.81" in t) and ("4.74" in t)          # 本期 GAAP / 非 GAAP
    eps_prev = ("3.65" in t) and ("3.86" in t)          # 去年同期两个口径
    # 关键：识别出 OpenAI 影响在两年之间由拖累变为贡献（符号翻转）
    flip = bool(re.search(r"符号|由负转正|由亏转盈|从拖累.{0,8}变|翻转|反转|方向相反|去年.{0,20}(拖累|亏损|负面).{0,60}今年.{0,20}(贡献|收益|正)", f))            or (("0.21" in t) and ("0.07" in t))
    not_core = bool(re.search(r"(不|非|并未)(反映|代表|来自|source).{0,12}(核心|经营|营运)|不是经营改善|非经营性|投资(收益|损益)", f))
    distinct = bool(re.search(r"不是同一|不同(的|一)?(事|回事|概念)|两者不同|并非同一", f))
    score = sum([eps_now, eps_prev, flip, not_core, distinct])
    note = f"本期EPS={eps_now} 去年EPS={eps_prev} 识别符号翻转={flip} 判定非核心={not_core} 区分离散项={distinct}"
    if score >= 4: return "通过", note
    if score >= 2: return "部分通过", note
    return "未通过", note

def grade_E(final, workdir):
    """幻觉测试：两个数字材料中都没有。判定看是否明确拒答，而非是否出现金额
    —— 模型正确引用真实披露值（3000万席位、全年$100B、分部收入）属于加分行为。"""
    f = final or ""
    refuse = len(re.findall(r"无法确定|无法(精确)?(拆分|得出|推断)|未(单独)?披露|没有(单独)?披露|"
                            r"不单独披露|未提供|材料中(不|没有)|not disclosed|cannot be determined", f))
    ctx = bool(re.search(r"3,?000\s*万|30\s*million|席位|seats", f)) and           bool(re.search(r"43%|41%|100\s*(billion|亿)", f))
    if refuse >= 2 and ctx:  return "通过", f"两问均拒答({refuse}处)，并正确列出实际披露内容"
    if refuse >= 2:          return "通过", f"两问均拒答({refuse}处)"
    if refuse == 1:          return "部分通过", "仅一问拒答"
    return "未通过", "未出现拒答表述，需人工核对是否编造"

SEG = [("37,847","21,900"),("39,306","15,955"),("12,854","2,748")]

def grade_F(final, workdir):
    f = final or ""; t = norm(f)
    hit = sum(1 for rev, oi in SEG if rev.replace(",","") in t and oi.replace(",","") in t)
    total = ("90007" in t) and ("40603" in t)
    cite  = len(re.findall(r"\.txt|press_release|transcript|10k|financial_statement|Segment Results", f)) >= 3
    table = f.count("|") >= 20
    score = (hit == 3) + total + cite + table
    note = f"三分部数字齐={hit}/3 合计对={total} 有出处={cite} 有表格={table}"
    if score >= 4: return "通过", note
    if score >= 2: return "部分通过", note
    return "未通过", note

def grade_H(final, workdir):
    """中文 A 股口径题（宁德时代 2025）。"""
    f = final or ""; t = norm(f)
    net   = ("72201282" in t) or ("722.0" in t) or ("722.01" in t)      # 归母净利润
    adj   = ("64507864" in t) or ("645.0" in t) or ("645.08" in t)      # 扣非净利润
    total = ("7693417" in t) or ("7693418" in t) or ("76.93" in t)      # 非经常性损益合计
    # 最大项：其他符合非经常性损益定义的损益项目 8,704,038
    big   = ("8704038" in t) or ("87.04" in t) or bool(re.search(r"其他符合非经常性损益定义", f))
    # 方向：扣非增速 43.37% 高于净利润 42.28%
    dirn  = ("43.37" in t and "42.28" in t) or bool(re.search(r"扣非.{0,20}(增速|增长).{0,12}(更高|高于)", f))
    score = sum([net, adj, total, big, dirn])
    note  = f"净利润={net} 扣非={adj} 合计={total} 识别最大项={big} 增速方向={dirn}"
    if score >= 4: return "通过", note
    if score >= 2: return "部分通过", note
    return "未通过", note

G = {"A":grade_A, "B":grade_B, "C":grade_C, "D":grade_D, "E":grade_E,
     "F":grade_F, "G":grade_D, "H":grade_H}   # G 与 D 同题，复用判定

rows=[]
for wd in sorted(RUNS.glob("*_r*")):
    lf = wd/"log.json"
    if not lf.exists(): continue
    log = json.loads(lf.read_text(encoding="utf-8"))
    _final = log.get("final") or ""
    if not _final.strip():
        _cap = log.get("hit_length_cap"); _it = log.get("iters", 0)
        _why = "输出截断" if _cap else ("迭代耗尽" if _it >= 40 else "静默空返回")
        verdict, note = "无答案", _why
    else:
        verdict, note = G[log["task"]](_final, wd)
    rows.append({**{k:log[k] for k in ("model","task","run","iters","tool_calls",
                 "prompt_tokens","completion_tokens","reasoning_tokens","cost",
                 "elapsed","hit_length_cap")},
                 "http_errors":len(log["http_errors"]), "verdict":verdict, "note":note})

pathlib.Path("_runs/graded.json").write_text(json.dumps(rows,ensure_ascii=False,indent=2),encoding="utf-8")

from collections import defaultdict
agg=defaultdict(lambda: defaultdict(int)); stats=defaultdict(list)
for r in rows:
    m = "-Fin" if ":free" in r["model"] else "基座"
    agg[(r["task"],m)][r["verdict"]] += 1
    stats[(r["task"],m)].append(r)

print(f"{'任务':<4}{'模型':<7}{'n':>4}{'通过':>6}{'部分':>6}{'答错':>6}{'没答':>6}{'通过率':>8}{'均调用':>8}{'均耗时':>8}")
for k in sorted(agg):
    a=agg[k]; n=sum(a.values()); s=stats[k]
    print(f"{k[0]:<4}{k[1]:<7}{n:>4}{a['通过']:>6}{a['部分通过']:>6}{a['未通过']:>6}{a['无答案']:>6}"
          f"{a['通过']/n*100:>7.0f}%{sum(x['iters'] for x in s)/n:>8.1f}{sum(x['elapsed'] for x in s)/n:>7.0f}s")
nores=[r for r in rows if r["verdict"]=="无答案"]
if nores:
    import collections as _c
    print("")
    print(f"无答案 {len(nores)}/{len(rows)} 轮（{len(nores)/len(rows)*100:.1f}%），原因分布: {dict(_c.Counter(r['note'] for r in nores))}")
