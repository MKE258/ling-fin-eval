"""把微软材料抽成纯文本，供模型输入。输出到 tests/_context/。"""
import sys, pathlib, re
sys.stdout.reconfigure(encoding="utf-8")
import openpyxl, docx

MAT = pathlib.Path(__file__).resolve().parent.parent / "materials"
OUT = pathlib.Path(__file__).resolve().parent / "_context"
OUT.mkdir(exist_ok=True)

def xlsx_to_text(path, sheets=None):
    wb = openpyxl.load_workbook(path, data_only=True)
    parts = []
    for name in (sheets or wb.sheetnames):
        ws = wb[name]
        parts.append(f"\n### 工作表: {name}")
        for row in ws.iter_rows(values_only=True):
            cells = ["" if v is None else (f"{v:g}" if isinstance(v, float) and v == int(v) else str(v)) for v in row]
            line = " | ".join(cells).rstrip(" |")
            if line.strip():
                parts.append(line)
    return "\n".join(parts)

def docx_to_text(path):
    d = docx.Document(path)
    parts = [p.text for p in d.paragraphs if p.text.strip()]
    for t in d.tables:
        for r in t.rows:
            cells = [c.text.strip() for c in r.cells]
            if any(cells):
                parts.append(" | ".join(cells))
    return re.sub(r"\n{3,}", "\n\n", "\n".join(parts))

jobs = {
    "financial_statements.txt": lambda: xlsx_to_text(MAT / "FinancialStatementFY26Q4.xlsx"),
    "metrics.txt":              lambda: xlsx_to_text(MAT / "Metrics_FY26Q4.xlsx"),
    "press_release.txt":        lambda: docx_to_text(MAT / "PressReleaseFY26Q4.docx"),
    "transcript.txt":           lambda: docx_to_text(MAT / "TranscriptFY26Q4.docx"),
    "10k.txt":                  lambda: docx_to_text(MAT / "MSFT_FY26Q4_10K.docx"),
}
print(f"{'文件':28s} {'字符数':>10s} {'约 token':>10s}")
total = 0
for name, fn in jobs.items():
    txt = fn()
    (OUT / name).write_text(txt, encoding="utf-8")
    total += len(txt)
    print(f"{name:28s} {len(txt):10,d} {len(txt)//3:10,d}")
print(f"{'合计':28s} {total:10,d} {total//3:10,d}   (context 上限 262,144)")
